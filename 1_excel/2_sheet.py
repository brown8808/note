
from openpyxl import Workbook # 새 워크북 생성
wb = Workbook() 
# wb.active
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = "Mysheet" #sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # rgb 형태로 값을 넣어주면 탭 색상 변경

# sheet, mysheet, yoursheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번째 index 에 sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet 에 접근

print(wb.sheetnames) # 현재 생성되어 있는 시트 이름 출력

 # Sheet 복사
new_ws["A1"] = "Test" # 뉴 워크시트 a1셀에 test 텍스트 입력
target = wb.copy_worksheet(new_ws) # 이걸 복사해서
target.title = "Copied Sheet" # 카피드 Sheet 생성 후 붙여넣기

wb.save("sample.xlsx") # 샘플로 저장
