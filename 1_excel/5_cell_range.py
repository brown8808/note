# -*- coding: utf-8 -*-
# 터미널 비우기 : cls
# 한번에 주석처리 : 컨트롤 + /

from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active



# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

 

col_B = ws["B"] # 영어 column 만 가지고 오기
#print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] # 영어, 수학 column 을 함께 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] # 첫번째 row 만 가지고 오기
# for cell in row_title:
#     print(cell.value)

#row_range = ws[2:6] # 2번째 줄에서 6번쨰 줄까지 가지고 오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate) # A10, az250
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy)
    
# 전체 rows
# print(tuple(ws.rows))   

# for row in tuple(ws.rows):
#     print(row[0].value)

# 전체 cloumns
# print(tuple(ws.columns))

# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): # 전체 row를 반복하며 가져오기
#     print(row[1].value)

# for column in ws.iter_cols(): # 전체 column
#     print(column[0].value)


# 2번째 줄부터 11번쨰 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value) # 수학, 영어


    
wb.save("sample.xlsx")

